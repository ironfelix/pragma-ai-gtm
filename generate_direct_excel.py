#!/usr/bin/env python3
"""
Генерация Excel для Яндекс.Директ в точном формате шаблона direct_example.xls
66 колонок, структура:
  [-] строки = ключевые слова (+ объявление 1)
  [+] строки = доп. объявления 2 и 3 для той же группы
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
ADGROUP_CSV  = BASE_DIR / "keywords-by-adgroup.csv"
NICHE_CSV    = BASE_DIR.parent / "pragma-ai-landing" / "semantics_ai_niche.csv"
OUT_FILE     = BASE_DIR / "k2_vnedrenie_direct.xlsx"

CAMPAIGN_NAME = "К2 Внедрение ИИ — Поиск"
REGION        = "Russia"
BID           = 10

# Минус-слова кампании (в формате Директа: через пробел с -)
CAMPAIGN_NEGATIVES = (
    "-бесплатно -бесплатный -скачать -своими руками -самостоятельно "
    "-курс -курсы -обучение -урок -туториал -гайд -диплом "
    "-вакансия -работа -зарплата -hh -headhunter "
    "-что такое -как работает -wikipedia "
    "-фото -картинки -видео -музыка -midjourney "
    "-форекс -акции -трейдинг -профессия -колледж -школа -вуз"
)

# ─── Объявления ───────────────────────────────────────────────────────────
GROUPS = {
    "Внедрение ИИ — общее": {
        "url":   "https://pragma-ai.ru/",
        "minus": "-бесплатно -курсы -обучение -скачать -вакансия",
        "ads": [
            ("Внедрение ИИ в бизнес — под ключ",  "Запуск 30 дней. Фикс цена.",       "AI Аудит 40 000 ₽: найдём точки автоматизации и рассчитаем ROI до старта."),
            ("Внедрение ИИ в компанию — 30 дней",  "Не интегратор за 2,5 млн.",        "Фикс цена в договоре. Интеграция с вашей CRM. Без скрытых доплат."),
            ("Разработка и внедрение ИИ — фикс",   "Результат за 30 дней. Фикс цена.", "Проектируем, разрабатываем, запускаем ИИ. 3 мес. поддержки в подарок."),
        ],
    },
    "Покупка / цена": {
        "url":   "https://pragma-ai.ru/",
        "minus": "-бесплатно -курсы -обучение -скачать -вакансия",
        "ads": [
            ("Внедрение ИИ под ключ — цена фикс",  "Старт от 40 000 ₽. 30 дней.",     "AI Аудит 40 000 ₽: находим точки автоматизации, считаем ROI до старта."),
            ("Сколько стоит внедрение ИИ?",         "Рассчитаем бесплатно. 30 минут.", "Разберём вашу задачу и назовём цену до начала работ. Без обязательств."),
            ("ИИ в бизнес под ключ — 30 дней",      "Цена фиксируется в договоре.",    "Никаких доплат по ходу проекта. Сроки и стоимость — в договоре."),
        ],
    },
    "ИИ в продажах и CRM": {
        "url":   "https://pragma-ai.ru/",
        "minus": "-бесплатно -курсы -обучение -скачать -вакансия",
        "ads": [
            ("ИИ для автоматизации продаж — 30 д",  "AmoCRM / Битрикс24. Фикс цена.",  "ИИ заполняет CRM после звонка. Автоматизируем воронку под ваш процесс."),
            ("AI автоматизация бизнеса под ключ",   "Запуск 14–30 дней. Фикс цена.",   "Автоматизируем поддержку, продажи, заполнение CRM. Без затяжных проектов."),
            ("AI-агент для AmoCRM — под ключ",       "Заполняет CRM сам. 30 дней.",     "Настраиваем ИИ-агента под вашу воронку продаж в AmoCRM или Битрикс24."),
        ],
    },
    "ИИ агенты": {
        "url":   "https://pragma-ai.ru/",
        "minus": "-бесплатно -курсы -обучение -конструктор -no-code -вакансия",
        "ads": [
            ("ИИ-агенты для бизнеса — под ключ",   "Старт 14–30 дней. Фикс цена.",    "ИИ-агент отвечает клиентам, квалифицирует лиды, заполняет CRM. 24/7."),
            ("Заказать ИИ-агента для бизнеса",      "Разработка 30 дней. Цена фикс.",  "Проектируем агента под ваши задачи. Интеграция с CRM, мессенджерами, 1С."),
            ("Голосовой ИИ-агент для бизнеса",      "Работает 24/7. Старт 30 дней.",   "Отвечает на вопросы, записывает, переключает. Не молчит, не устаёт."),
        ],
    },
    "HR автоматизация": {
        "url":   "https://pragma-ai.ru/",
        "minus": "-бесплатно -курсы -написать резюме -пример резюме -hh -superjob -вакансия",
        "ads": [
            ("Скрининг резюме — ИИ-ассистент",       "100 резюме за 10 минут. Фикс цена", "ИИ анализирует резюме, задаёт вопросы. Рекрутер видит только подходящих."),
            ("Автоматизация подбора персонала AI",   "Цикл найма — втрое короче.",        "HR-агент 24/7: скрининг, первичное интервью, рейтинг кандидатов."),
            ("Бот для скрининга резюме — ИИ",        "Без ручного разбора. Фикс цена.",   "ИИ проводит первичные интервью 24/7. Рекрутер тратит время только на финал."),
        ],
    },
}

NICHE_ADDITIONS = {
    "Внедрение ИИ — общее": ["ии для бизнеса", "внедрить ии", "внедрить ии в бизнес", "внедрение ai", "ai решения для бизнеса"],
    "Покупка / цена":        ["купить ии"],
    "ИИ агенты":             ["ai агент", "ai агенты", "ии агент на заказ", "ии агент для сайта"],
}

# ─── Загрузка ключей ───────────────────────────────────────────────────────
def load_keywords(phase1_only=True):
    keys_by_group = {g: [] for g in GROUPS}

    with open(ADGROUP_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if not row["campaign"].startswith("К2_"):
                continue
            group = row["group"]
            if group not in keys_by_group:
                continue
            avbid = row["avbid"].strip()
            if phase1_only and avbid not in ("0", ""):
                continue
            kw = row["keyword"].strip()
            if kw:
                keys_by_group[group].append(kw)

    with open(NICHE_CSV, encoding="utf-8-sig") as f:
        niche_map = {row["word"]: row for row in csv.DictReader(f)}

    for group, words in NICHE_ADDITIONS.items():
        existing = set(keys_by_group[group])
        for w in words:
            if w in niche_map:
                avbid = niche_map[w]["avbid"].strip()
                if phase1_only and avbid not in ("0", ""):
                    continue
            if w not in existing:
                keys_by_group[group].append(w)

    return keys_by_group

# ─── Генерация Excel ───────────────────────────────────────────────────────
def make_row(n_cols=66):
    return [""] * n_cols

def generate(phase1_only=True):
    keys_by_group = load_keywords(phase1_only)

    wb = Workbook()
    ws = wb.active
    ws.title = "Texts"

    N = 66  # всего колонок

    # ── Заголовок шаблона (строки 1-9) ──
    ws.append(make_row(N))
    ws.append(make_row(N))
    ws.append(make_row(N))
    ws.append(make_row(N))
    ws.append(make_row(N))

    meta = make_row(N)
    meta[0] = "Pragma AI — К2 Внедрение ИИ"
    ws.append(meta)

    ws.append(make_row(N))

    order = make_row(N)
    order[0] = "Order No.:"
    order[1] = "pragma-k2"
    order[2] = "Currency:"
    order[3] = "RUB"
    ws.append(order)

    ws.append(make_row(N))

    # ── Строка заголовков (строка 10) ──
    headers = make_row(N)
    headers[0]  = "Additional group ad"
    headers[1]  = "Ad type"
    headers[2]  = "Group ID"
    headers[3]  = "Group name"
    headers[4]  = "Group number"
    headers[5]  = "Keyword ID"
    headers[6]  = "Keyword (with negative keywords)"
    headers[7]  = "Ad ID"
    headers[8]  = "Title 1"
    headers[9]  = "Title 2"
    headers[10] = "Ad Text"
    headers[11] = "Length"
    headers[46] = "Link"
    headers[47] = "Display link"
    headers[48] = "Region"
    headers[49] = "Company from Yandex Business"
    headers[50] = "Bid"
    headers[51] = "Bid in ad networks"
    headers[52] = "Ad status"
    headers[53] = "Keyword status"
    headers[54] = "Sitelink titles"
    headers[55] = "Sitelink descriptions"
    headers[56] = "Sitelink URLs"
    headers[57] = "Parameter 1"
    headers[58] = "Parameter 2"
    headers[59] = "Labels"
    headers[60] = "Image"
    headers[61] = "Creative"
    headers[62] = "Moderation status of creative"
    headers[63] = "Callouts"
    headers[64] = "Negative keywords in group"
    headers[65] = "Age rating"

    # Стиль заголовков
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, N + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 35

    # ── Пустая строка 11 (под-заголовок комбинаторных — оставляем пустой) ──
    ws.append(make_row(N))

    # ── Данные ──
    group_colors = {
        "Внедрение ИИ — общее": "D6E4F0",
        "Покупка / цена":        "D5F0D6",
        "ИИ в продажах и CRM":   "FFF3CD",
        "ИИ агенты":             "F8D7DA",
        "HR автоматизация":      "E8D5F0",
    }

    group_num = 1
    total_rows = 0

    for group_name, group_data in GROUPS.items():
        keywords = keys_by_group.get(group_name, [])
        if not keywords:
            group_num += 1
            continue

        fill_hex = group_colors.get(group_name, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_hex)
        ads = group_data["ads"]
        h1_0, h2_0, text_0 = ads[0]  # первое объявление — идёт с ключами

        # [-] строки: по одной на каждый ключ, с первым объявлением
        for kw_idx, keyword in enumerate(keywords):
            row = make_row(N)
            row[0]  = "-"
            row[1]  = "Text & Image"
            row[3]  = group_name
            row[4]  = str(group_num)
            row[6]  = keyword
            row[8]  = h1_0
            row[9]  = h2_0
            row[10] = text_0
            row[46] = group_data["url"]
            row[48] = REGION
            row[50] = BID
            # Минус-слова только в первой строке группы
            if kw_idx == 0:
                row[64] = group_data["minus"]
            ws.append(row)

            # Цвет строки
            cur_row = ws.max_row
            for col in range(1, N + 1):
                ws.cell(row=cur_row, column=col).fill = fill
            total_rows += 1

        # [+] строки: доп. объявления 2 и 3 (без ключей)
        for ad_idx in range(1, len(ads)):
            h1, h2, text = ads[ad_idx]
            row = make_row(N)
            row[0]  = "+"
            row[1]  = "Text & Image"
            row[3]  = group_name
            row[4]  = str(group_num)
            row[8]  = h1
            row[9]  = h2
            row[10] = text
            row[46] = group_data["url"]
            ws.append(row)

            cur_row = ws.max_row
            for col in range(1, N + 1):
                ws.cell(row=cur_row, column=col).fill = fill
            total_rows += 1

        group_num += 1

    # ── Ширина колонок ──
    col_widths = {
        1: 6,   # +/-
        2: 14,  # Ad type
        4: 25,  # Group name
        5: 8,   # Group number
        7: 45,  # Keyword
        9: 42,  # Title 1
        10: 32, # Title 2
        11: 60, # Ad text
        47: 40, # Link
        49: 15, # Region
        51: 8,  # Bid
        65: 45, # Negatives
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A12"

    wb.save(OUT_FILE)
    return keys_by_group, total_rows

# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    print("Генерируем Excel в формате Яндекс.Директ...")
    keys, total = generate(phase1_only=True)

    kw_total = sum(len(v) for v in keys.values())
    print(f"\nФайл:     {OUT_FILE.name}")
    print(f"Ключей:   {kw_total} (только avbid=0)")
    print(f"Строк:    {total}")
    print()
    for g, kws in keys.items():
        print(f"  {g}: {len(kws)} ключей + 2 доп. объявления")
    print(f"\nОткрываем файл...")

if __name__ == "__main__":
    main()
